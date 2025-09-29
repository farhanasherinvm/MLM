from rest_framework import generics, permissions
from .models import Profile
from .serializers import ProfileSerializer
from rest_framework.response import Response
from .models import KYC
from .serializers import KYCSerializer
from rest_framework.views import APIView
from .serializers import ReferralSerializer
from django.db.models import Q
from django.contrib.auth import get_user_model
from datetime import datetime, timedelta


from rest_framework.exceptions import PermissionDenied

from rest_framework.permissions import IsAuthenticated

from .serializers import ReferralListSerializer,CurrentUserProfileSerializer
from .utils import get_all_referrals
from rest_framework.permissions import IsAdminUser
from django.utils.dateparse import parse_date

from django.utils.timezone import make_aware, is_naive
from datetime import datetime
from django.http import HttpResponse
import csv
from io import BytesIO

from django.utils.timezone import is_naive, make_aware

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4


CustomUser = get_user_model()
from django.db import models

from django.db.models import Value
from django.db.models.functions import Concat

from rest_framework.pagination import PageNumberPagination




class ProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = ProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        # Fetch profile of currently logged-in user
        return self.request.user.profile


class KYCView(generics.RetrieveUpdateAPIView):
    serializer_class = KYCSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        # Always get the existing KYC for the user
        try:
            obj = KYC.objects.get(user=self.request.user)
        except KYC.DoesNotExist:
            # Allow creating only if user doesn't have one
            obj = KYC(user=self.request.user)
        return obj

    def update(self, request, *args, **kwargs):
        obj = self.get_object()

        # Users can only update once; after verified, only admin can update
        if obj.pk is not None and not request.user.is_staff:
            raise PermissionDenied("You cannot update KYC again. Please contact admin.")

        # If obj exists in DB, update; if it's new, save first
        return super().update(request, *args, **kwargs)


    def perform_update(self, serializer):
        # Save serializer normally
        serializer.save()

class ReferralView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        referral_id = user.user_id  

        serializer = ReferralSerializer(data={"referral_id": referral_id})
        serializer.is_valid(raise_exception=True)

        return Response(serializer.data)


class ReferralPagination(PageNumberPagination):
    page_size = 12  # items per page
    page_size_query_param = 'limit'  # optional override via ?limit=
    max_page_size = 50

class ReferralListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        all_referrals = get_all_referrals(user, max_level=6)

        #  Query params 
        status = request.query_params.get("status")
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")  # 'csv', 'pdf', 'xlsx'
        user_id = request.query_params.get("user_id")
        fullname = request.query_params.get("fullname")
        email = request.query_params.get("email")
        mobile = request.query_params.get("mobile")
        fromdate = request.query_params.get("fromdate") or request.query_params.get("from_date")
        enddate = request.query_params.get("enddate") or request.query_params.get("end_date")
        referred_by_id = request.query_params.get("referred_by_id")
        referred_by_name = request.query_params.get("referred_by_name")

        # ---------------- Filters ----------------
        if status and status.lower() != "all":
            if status.lower() == "active":
                all_referrals = [r for r in all_referrals if r.is_active]
            elif status.lower() == "inactive":
                all_referrals = [r for r in all_referrals if not r.is_active]

        if user_id:
            all_referrals = [r for r in all_referrals if str(r.user_id).lower() == user_id.lower()]

        if fullname:
            fullname_lower = fullname.lower()
            all_referrals = [
                r for r in all_referrals
                if f"{r.first_name} {r.last_name}".strip().lower().find(fullname_lower) != -1
            ]

        if email:
            email_lower = email.lower()
            all_referrals = [r for r in all_referrals if r.email and email_lower in r.email.lower()]

        if mobile:
            all_referrals = [r for r in all_referrals if r.mobile and mobile in r.mobile]

        if referred_by_id:
            all_referrals = [
                r for r in all_referrals
                if r.sponsor_id and str(getattr(r.sponsor_id, "user_id", r.sponsor_id)) == str(referred_by_id)
            ]

        if referred_by_name:
            referred_by_name_lower = referred_by_name.strip().lower()
            all_referrals = [
                r for r in all_referrals
                if r.sponsor_id and (
                    (hasattr(r.sponsor_id, "first_name") and
                     f"{r.sponsor_id.first_name} {r.sponsor_id.last_name}".strip().lower() == referred_by_name_lower)
                    or CustomUser.objects.annotate(
                        fullname=Concat("first_name", Value(" "), "last_name")
                    ).filter(
                        user_id=r.sponsor_id,
                        fullname__iexact=referred_by_name
                    ).exists()
                )
            ]

        # Date range filter 
        if fromdate or enddate:
            try:
                from_dt = datetime.strptime(fromdate, "%Y-%m-%d") if fromdate else datetime.min
                end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) if enddate else datetime.max
                if is_naive(from_dt):
                    from_dt = make_aware(from_dt)
                if is_naive(end_dt):
                    end_dt = make_aware(end_dt)
                all_referrals = [
                    r for r in all_referrals
                    if r.date_of_joining and from_dt <= r.date_of_joining <= end_dt
                ]
            except ValueError:
                pass

        # Sorting by joining date 
        def get_joined_date(u):
            if u.date_of_joining:
                dt = u.date_of_joining
                if is_naive(dt):
                    dt = make_aware(dt)
                return dt
            return make_aware(datetime.min)

        all_referrals.sort(key=get_joined_date, reverse=True)

        #  Limit (optional)
        if limit:
            try:
                limit = int(limit)
                all_referrals = all_referrals[:limit]
            except ValueError:
                pass

        #  Pagination 
        paginator = ReferralPagination()
        page = paginator.paginate_queryset(all_referrals, request, view=self)

        # Serializer 
        serializer = ReferralListSerializer(page, many=True)

        # Prepare export data 
        data_rows = []
        for r in serializer.data:
            full_name = r.get('fullname', f"{r.get('first_name','')} {r.get('last_name','')}".strip())
            data_rows.append([
                r.get('user_id', ''),
                full_name,
                r.get('email', ''),
                r.get('mobile', ''),
                r.get('joined_date', ''),
                r.get('total_count', 0),
                r.get('level', ''),
                r.get('status', ''),
            ])

        headings = ["User ID", "Full Name", "Email", "Mobile", "Date of Joining", "Referral Count", "Rank", "Status"]

        # Export 
        if export == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="referrals_export.csv"'
            writer = csv.writer(response)
            writer.writerow(headings)
            writer.writerows(data_rows)
            return response

        elif export == "pdf":
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            y = height - 50
            p.setFont("Helvetica-Bold", 14)
            p.drawString(200, y, "Referral Export")
            y -= 30
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, y, " | ".join(headings))
            y -= 20
            p.setFont("Helvetica", 9)
            for row in data_rows:
                line = " | ".join(str(item) for item in row)
                p.drawString(50, y, line)
                y -= 15
                if y < 50:
                    p.showPage()
                    y = height - 50
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(50, y, " | ".join(headings))
                    y -= 20
                    p.setFont("Helvetica", 9)
            p.save()
            pdf = buffer.getvalue()
            buffer.close()
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="referrals_export.pdf"'
            return response

        elif export == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Referrals Export"
            ws.append(headings)
            for row in data_rows:
                ws.append(row)
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="referrals_export.xlsx"'
            return response

        #  Return paginated JSON response
        return paginator.get_paginated_response(serializer.data)


class AdminHomeView(APIView):
    permission_classes = [IsAdminUser]  
    def get(self, request):
        total_users = CustomUser.objects.count()

        data = {
            "total_users": total_users,
        }
        return Response(data)



class ReferralExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        all_referrals = get_all_referrals(user, max_level=6)

        # ---------------- Query params ----------------
        status = request.query_params.get("status")
        email = request.query_params.get("email")
        fullname = request.query_params.get("fullname")
        mobile = request.query_params.get("mobile")
        user_id = request.query_params.get("user_id")
        from_date = request.query_params.get("from_date")
        end_date = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")  # 'csv', 'pdf', 'xlsx'

        #  Filters 
        if status and status.lower() != "all":
            all_referrals = [r for r in all_referrals if (r.is_active if status.lower() == "active" else not r.is_active)]

        if email:
            all_referrals = [r for r in all_referrals if r.email and email.lower() in r.email.lower()]

        if fullname:
            all_referrals = [
                r for r in all_referrals
                if fullname.lower() in f"{r.first_name} {r.last_name}".lower()
            ]

        if mobile:
            all_referrals = [r for r in all_referrals if r.mobile and mobile in r.mobile]

        if user_id:
            all_referrals = [r for r in all_referrals if r.user_id and user_id.lower() in r.user_id.lower()]

        if from_date:
            try:
                from_dt = datetime.strptime(from_date, "%Y-%m-%d")
                all_referrals = [
                    r for r in all_referrals
                    if r.date_of_joining and r.date_of_joining.date() >= from_dt.date()
                ]
            except ValueError:
                pass

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                all_referrals = [
                    r for r in all_referrals
                    if r.date_of_joining and r.date_of_joining.date() <= end_dt.date()
                ]
            except ValueError:
                pass

        #  Sort by joining date 
        def get_joined_date(u):
            if u.date_of_joining:
                dt = u.date_of_joining
                if is_naive(dt):
                    dt = make_aware(dt)
                return dt
            return datetime.min.replace(tzinfo=None)

        all_referrals.sort(key=get_joined_date, reverse=True)

        # Limit 
        if limit:
            try:
                limit = int(limit)
                all_referrals = all_referrals[:limit]
            except ValueError:
                pass

        #  Pagination
        paginator = ReferralPagination()
        page = paginator.paginate_queryset(all_referrals, request, view=self)
        serializer = ReferralListSerializer(page, many=True)

        #  Prepare Data Rows 
        data_rows = []
        for r in serializer.data:
            full_name = r.get('fullname', f"{r.get('first_name','')} {r.get('last_name','')}".strip())
            data_rows.append([
                full_name,
                r.get('email', ''),
                r.get('mobile', ''),
                r.get('user_id', ''),
                r.get('position', ''),
                r.get('referred_by_name', ''),
                r.get('joined_date', ''),
                r.get('status', ''),
            ])

        headings = ["Full Name", "Email", "Mobile", "User ID", "Position", "Referred By", "Joining Date", "Status"]

        # Export 
        if export == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="referrals_export.csv"'
            writer = csv.writer(response)
            writer.writerow(headings)
            writer.writerows(data_rows)
            return response

        elif export == "pdf":
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            y = height - 50
            p.setFont("Helvetica-Bold", 14)
            p.drawString(150, y, "Referral Export")
            y -= 30
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, y, " | ".join(headings))
            y -= 20
            p.setFont("Helvetica", 10)
            for row in data_rows:
                line = " | ".join(str(item) for item in row)
                p.drawString(50, y, line)
                y -= 20
                if y < 50:
                    p.showPage()
                    y = height - 50
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(50, y, " | ".join(headings))
                    y -= 20
                    p.setFont("Helvetica", 10)
            p.save()
            pdf = buffer.getvalue()
            buffer.close()
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="referrals_export.pdf"'
            return response

        elif export == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Referrals Export"
            ws.append(headings)
            for row in data_rows:
                ws.append(row)
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="referrals_export.xlsx"'
            return response

        #  Default JSON Response 
        return paginator.get_paginated_response(serializer.data)

class CurrentUserProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            profile = Profile.objects.get(user=request.user)
        except Profile.DoesNotExist:
            return Response({"detail": "Profile not found."}, status=404)

        serializer = CurrentUserProfileSerializer(profile)
        return Response(serializer.data)