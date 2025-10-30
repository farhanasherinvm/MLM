from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAdminUser
from django.db.models import Q, Sum, Count, F,IntegerField, DecimalField
from django.utils import timezone
from level.models import UserLevel, LevelPayment,Level
from users.models import CustomUser
from .models import AdminNotification
from .serializers import (
    AUCReportSerializer,
    AdminNotificationSerializer,
    AdminSendRequestReportSerializer,
    AdminPaymentSerializer,
    AdminSummaryAnalyticsSerializer,
    UserAnalyticsSerializer
) # Only import the specific serializers
from django.http import HttpResponse
import csv
import io
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from openpyxl import Workbook
import logging
from rest_framework.generics import ListAPIView
from rest_framework import status
from decimal import Decimal
from django.db.utils import OperationalError, ProgrammingError
from django.core.exceptions import FieldError
from django.db.models.expressions import OuterRef, Subquery 
from users.models import Payment
from level.models import PmfPayment
from operator import attrgetter
from decimal import Decimal
from datetime import timedelta
from django.utils import timezone

logger = logging.getLogger(__name__)

# Helper function for safe date conversion (used in both views)
def safe_parse_date(date_str):
    try:
        return timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


class AdminAUCReportView(APIView):
    permission_classes = [IsAdminUser]
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10
    
    def get_all_payment_data(self):
        # 1. Fetch Registration Payments (must be Verified)
        registration_payments = Payment.objects.filter(
            status='Verified' 
        ).select_related('user').all()

        # 2. Fetch PMF Payments (must be Verified)
        pmf_payments = PmfPayment.objects.filter(
            status='Verified'
        ).select_related('user').all()

        # Combine and sort by creation date (descending)
        combined_data = sorted(
            list(registration_payments) + list(pmf_payments),
            key=attrgetter('created_at'),
            reverse=True
        )
        return combined_data

    def apply_filters(self, request, combined_data):
        

        email = request.query_params.get("email", '').lower()
        user_id = request.query_params.get("user_id", '').lower()
        search = request.query_params.get('search', '').lower()
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        
        start_date = safe_parse_date(start_date_str)
        end_date = safe_parse_date(end_date_str)
        
        filtered_data = []

        for item in combined_data:
            # Need to re-implement the user retrieval utility here if it's not a method on the view
            def _get_user_from_item(obj):
                if hasattr(obj, 'user') and obj.user:
                    return obj.user
                
                if obj.__class__.__name__ == 'Payment':
                    payment_user_id = getattr(obj, 'user_id', None)
                    
                    if payment_user_id:
                        try:
                            return CustomUser.objects.get(user_id=payment_user_id)
                        except CustomUser.DoesNotExist:
                            pass
                            
                return None
            
            user = _get_user_from_item(item)

            if user is None:
                continue 
            
            # --- Text/ID Filters ---
            match_text = True
            if search or user_id or email:
                user_match = False
                if user:
                    if user_id and user.user_id.lower() == user_id:
                        user_match = True
                    elif email and user.email and user.email.lower().find(email) != -1:
                        user_match = True
                    elif search:
                        name = f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".lower()
                        if user.user_id.lower().find(search) != -1 or name.find(search) != -1 or (user.email and user.email.lower().find(search) != -1):
                            user_match = True
                
                if (user_id or email or search) and not user_match:
                    match_text = False

            # --- Date Filters ---
            match_date = True
            item_date = item.created_at.date() if item.created_at else None
            
            if item_date:
                if start_date and item_date < start_date:
                    match_date = False
                if end_date and item_date > end_date:
                    match_date = False
            else:
                if start_date or end_date:
                    match_date = False

            if match_text and match_date:
                filtered_data.append(item)

            if limit:
                try:
                    limit_int = int(limit)
                    filtered_data = filtered_data[:limit_int]
                    # If we're exporting, we use this limited set.
                except ValueError:
                    pass 

        return filtered_data


    def get(self, request):
        combined_data = self.get_all_payment_data()
        
        # 1. Apply Date Range Logic and Prepare Query Params
        date_range = request.query_params.get("date_range", '').lower()
        today = timezone.localdate()

        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")

        if date_range == 'today':
            start_date_str = today.isoformat()
            end_date_str = today.isoformat()
        elif date_range == 'this_week':
            start_of_week = today - timedelta(days=today.weekday())
            start_date_str = start_of_week.isoformat()
            end_date_str = today.isoformat()
        elif date_range == 'this_month':
            start_of_month = today.replace(day=1)
            start_date_str = start_of_month.isoformat()
            end_date_str = today.isoformat()
        elif date_range == 'this_year':
            start_of_year = today.replace(month=1, day=1)
            start_date_str = start_of_year.isoformat()
            end_date_str = today.isoformat()

        # Update request query params so apply_filters can use the calculated dates
        if start_date_str is not None or end_date_str is not None:
            request.query_params._mutable = True
            request.query_params['start_date'] = start_date_str if start_date_str else ''
            request.query_params['end_date'] = end_date_str if end_date_str else ''
            request.query_params._mutable = False

        # Collect all filter parameters for export header
        filter_params = {
            'User ID': request.query_params.get("user_id", ''),
            'Search Term': request.query_params.get('search', ''),
            'Email': request.query_params.get("email", ''),
            'Start Date': start_date_str, 
            'End Date': end_date_str,
            'Date Range Key': date_range if date_range not in ('', None) else 'Manual'
        }
        # --------------------------------------------------------------------------------

        # 2. Apply Filters
        filtered_data = self.apply_filters(request, combined_data)
        
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")
        
        # 3. Handle Limit
        if limit:
            try:
                limit_int = int(limit)
                queryset_for_export = filtered_data[:limit_int]
            except ValueError:
                pass 

        # 4. Serialize all filtered data to calculate totals, regardless of pagination/limit
        full_serialized_data = AUCReportSerializer(filtered_data, many=True).data

        # 5. Calculate Total Payments and GST
        total_payments = sum((Decimal(item['amount']) for item in full_serialized_data), Decimal('0.00'))
        total_gst = sum((Decimal(item['gst_total']) for item in full_serialized_data), Decimal('0.00'))
        total_cgst = sum((Decimal(item['cgst']) for item in full_serialized_data), Decimal('0.00'))
        total_sgst = sum((Decimal(item['sgst']) for item in full_serialized_data), Decimal('0.00'))
            
        totals = {
            'total_payments': total_payments.quantize(Decimal('0.01')),
            'total_gst': total_gst.quantize(Decimal('0.01')),
            'total_cgst': total_cgst.quantize(Decimal('0.01')),
            'total_sgst': total_sgst.quantize(Decimal('0.01')),
        }

        # 6. Handle Export
        if export in ["csv", "pdf", "xlsx"]:
            # Use the filtered_data for export and pass filter_params
            if export == "csv":
                return self.export_csv(filtered_data, totals, 'auc_report', filter_params)
            elif export == "pdf":
                return self.export_pdf(filtered_data, totals, 'auc_report', filter_params)
            elif export == "xlsx":
                return self.export_xlsx(filtered_data, totals, 'auc_report', filter_params)

        # 7. Handle Pagination (for JSON response)
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(filtered_data, request)
        
        if page is not None:
            serializer = AUCReportSerializer(page, many=True)
            response_data = paginator.get_paginated_response(serializer.data).data
            # Add totals to the paginated response
            response_data['report_totals'] = totals
            return Response(response_data)
            
        # If no pagination/limit, return all data + totals
        return Response({
            'results': full_serialized_data,
            'report_totals': totals
        })

    # --------------------------------------------------------------------------------
    # --- Export Methods (Updated to include totals and filters) ---
    # --------------------------------------------------------------------------------

    def export_csv(self, queryset, totals, filename_prefix, filter_params):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        
        # --- ADD FILTER INFO ---
        writer.writerow(['Report Filters:'])
        for key, value in filter_params.items():
            if value:
                writer.writerow([f"{key}:", value])
        writer.writerow([])
        # -----------------------
        
        writer.writerow(['User ID', 'Name', 'Phone', 'Email', 'Type', 'Amount', 'Status', 'Date', 'GST Total', 'CGST', 'SGST'])
        
        serializer = AUCReportSerializer(queryset, many=True)
        for data in serializer.data:
            writer.writerow([
                data['user_id'], data['user_name'], data['phone_number'], data['email'], 
                data['transaction_type'], str(data['amount']), data['status'], 
                str(data['date']) if data['date'] else '', 
                str(data['gst_total']), str(data['cgst']), str(data['sgst'])
            ])
            
        # Add totals section
        writer.writerow([])
        writer.writerow(['TOTALS:', '', '', '', '', ''])
        writer.writerow(['Total Payments', str(totals['total_payments'])])
        writer.writerow(['Total GST', str(totals['total_gst'])])
        writer.writerow(['Total CGST', str(totals['total_cgst'])])
        writer.writerow(['Total SGST', str(totals['total_sgst'])])
        
        return response

    def export_pdf(self, queryset, totals, filename_prefix, filter_params):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"{filename_prefix.replace('_', ' ').upper()} REPORT", styles["Title"]))
        elements.append(Spacer(1, 12))

        # --- ADD FILTER INFO ---
        filter_text = []
        for key, value in filter_params.items():
            if value:
                filter_text.append(f"<b>{key}:</b> {value}")

        if filter_text:
            elements.append(Paragraph(" | ".join(filter_text), styles["BodyText"]))
            elements.append(Spacer(1, 12))
        # -----------------------
        
        # --- Data Table ---
        # 1. CORRECTED HEADER (10 Columns - Email removed)
        data = [['User ID', 'Name', 'Phone', 'Type', 'Amount', 'Status', 'Date', 'GST Total', 'CGST', 'SGST']]
        
        serializer = AUCReportSerializer(queryset, many=True)
        for data_item in serializer.data:
            # 2. CORRECTED DATA ROW (10 Items - Email removed)
            data.append([
                data_item['user_id'], 
                data_item['user_name'], 
                data_item['phone_number'],
                # data_item['email'], <-- REMOVED THIS LINE
                data_item['transaction_type'], 
                str(data_item['amount']), 
                data_item['status'], 
                str(data_item['date']) if data_item['date'] else '', 
                str(data_item['gst_total']), 
                str(data_item['cgst']),     
                str(data_item['sgst'])      
            ])

        table_width = A4[0] - 80 
        # 3. RECALCULATED ColWidths for 10 columns (Approximate widths for better fit)
        col_widths = [table_width * w for w in [0.10, 0.18, 0.12, 0.10, 0.08, 0.08, 0.12, 0.09, 0.06, 0.07]]
        table = Table(data, colWidths=col_widths)
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCCC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'), # Align Amount and all GST columns to the right (Index 4 is 'Amount')
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 24))
        
        # --- Totals Table ---
        totals_data = [
            ['Report Totals', ''],
            ['Total Payments', f"₹ {totals['total_payments']}"],
            ['Total GST (18%)', f"₹ {totals['total_gst']}"],
            ['Total CGST (9%)', f"₹ {totals['total_cgst']}"],
            ['Total SGST (9%)', f"₹ {totals['total_sgst']}"],
        ]
        
        totals_table = Table(totals_data, colWidths=[2.5*inch, 1.5*inch])
        totals_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(totals_table)

        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response


    def export_xlsx(self, queryset, totals, filename_prefix, filter_params):
        from openpyxl import Workbook 
        
        wb = Workbook()
        ws = wb.active
        ws.title = "AUCReportSimple"
        
        # --- ADD FILTER INFO ---
        ws.append(['Report Filters:'])
        row = 2
        for key, value in filter_params.items():
            if value:
                ws.cell(row=row, column=1, value=f"{key}:")
                ws.cell(row=row, column=2, value=value)
                row += 1
        
        ws.append([]) # Add a blank row for separation
        # -----------------------

        # Headers based on your example structure, excluding GST fields
        ws.append(['User ID', 'User Name', 'Phone', 'Type', 'Amount', 'Status', 'Date', 'Total Payment', 'Total GST', 'Total CGST', 'Total SGST'])
        
        serializer = AUCReportSerializer(queryset, many=True)
        for data in serializer.data:
            ws.append([
                data.get('user_id', 'N/A'),
                data.get('user_name', 'N/A'),
                data.get('phone_number', 'N/A'),
                data.get('transaction_type', 'N/A'),
                data.get('amount', 0.00),
                data.get('status', 'N/A'),
                data.get('date', ''),
                # Include all original fields for completeness, even if not in the simple header, 
                # or just use the fields you explicitly want in the header:
                data.get('gst_total', 0.00), 
                data.get('cgst', 0.00),
                data.get('sgst', 0.00)
            ])
            
        # Auto-size columns (Standard best practice for Excel exports)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
# ----------------------------------------------------------------------------
# --- NEW SPECIFIC REPORT VIEWS (Replacing AllAdminReportsView logic) ---
# ----------------------------------------------------------------------------

class AdminSendRequestView(APIView):
    permission_classes = [IsAdminUser]
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10

    def get(self, request):
        queryset = UserLevel.objects.select_related('user', 'level').prefetch_related('payments').all().order_by('-requested_date')

        email = request.query_params.get("email")
        status = request.query_params.get("status")
        user_id = request.query_params.get("user_id")
        username = request.query_params.get("username")
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")
        search = request.query_params.get('search', '')
        
        start_date = safe_parse_date(start_date_str)
        end_date = safe_parse_date(end_date_str)

        q_objects = Q()
        if email:
            q_objects &= Q(user__email__icontains=email)
        if user_id:
            q_objects &= Q(user__user_id__iexact=user_id)
        if username:
            q_objects &= (Q(user__first_name__icontains=username) | Q(user__last_name__icontains=username))
        
        if status and status.lower() != "all":
            if status.lower() == "completed":
                q_objects &= Q(status='paid')
            elif status.lower() == "pending":
                q_objects &= ~Q(status='paid')
        
        if search:
            q_objects &= (
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__user_id__icontains=search) |
                Q(status__icontains=search)
            )

        queryset = queryset.filter(q_objects)
        
        if start_date:
            queryset = queryset.filter(requested_date__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(requested_date__date__lte=end_date)
            
        if limit:
            try:
                limit = int(limit)
                queryset = queryset[:limit]
            except ValueError:
                pass
        
        if export:
            export_data = AdminSendRequestReportSerializer(queryset, many=True, context={'request': request}).data
            if export == "csv":
                return self._export_csv_send_request(export_data, 'send_request_report')
            elif export == "pdf":
                return self._export_pdf_send_request(export_data, 'send_request_report')
            elif export == "xlsx":
                return self._export_xlsx_send_request(export_data, 'send_request_report')

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        
        serializer = AdminSendRequestReportSerializer(page, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)

    def _export_csv_send_request(self, data, filename_prefix):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'From User', 'Username', 'From Name', 'Amount', 'Status', 'Requested Date', 'Payment Method', 'Level', 'Linked Username'])
        
        for item in data:
            writer.writerow([
                item['id'], item['from_user'], item['username'], item['from_name'], str(item['amount']),
                item['status'], str(item['requested_date']) if item['requested_date'] else '',
                item['payment_method'], item['level'], item['linked_username']
            ])
        return response

    def _export_pdf_send_request(self, data, filename_prefix):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"{filename_prefix.replace('_', ' ').title()} Report", styles["Title"]))

        table_data = [['ID', 'From User', 'Username', 'From Name', 'Amount', 'Status', 'Requested Date', 'Payment Method', 'Level', 'Linked Username']]
        for item in data:
            table_data.append([
                str(item['id']), item['from_user'], item['username'], item['from_name'], str(item['amount']),
                item['status'], str(item['requested_date']) if item['requested_date'] else '',
                item['payment_method'], item['level'], item['linked_username']
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _export_xlsx_send_request(self, data, filename_prefix):
        wb = Workbook()
        ws = wb.active
        ws.title = "SendRequestReport"
        ws.append(['ID', 'From User', 'Username', 'From Name', 'Amount', 'Status', 'Requested Date', 'Payment Method', 'Level', 'Linked Username'])
        for item in data:
            ws.append([
                item['id'], item['from_user'], item['username'], item['from_name'], item['amount'],
                item['status'], str(item['requested_date']) if item['requested_date'] else '',
                item['payment_method'], item['level'], item['linked_username']
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

class AdminPaymentReportView(APIView):
    permission_classes = [IsAdminUser]
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10

    def get(self, request):
        queryset = LevelPayment.objects.select_related('user_level__user', 'user_level__level').all().order_by('-created_at')

        email = request.query_params.get("email")
        status = request.query_params.get("status")
        user_id = request.query_params.get("user_id")
        username = request.query_params.get("username")
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")
        search = request.query_params.get('search', '')

        start_date = safe_parse_date(start_date_str)
        end_date = safe_parse_date(end_date_str)

        q_objects = Q()
        if email:
            q_objects &= Q(user_level__user__email__icontains=email)
        if user_id:
            q_objects &= Q(user_level__user__user_id__iexact=user_id)
        if username:
            q_objects &= (Q(user_level__user__first_name__icontains=username) | Q(user_level__user__last_name__icontains=username))
        
        if status and status.lower() != "all":
            q_objects &= Q(status__iexact=status)
        
        if search:
            q_objects &= (
                Q(user_level__user__first_name__icontains=search) |
                Q(user_level__user__last_name__icontains=search) |
                Q(user_level__user__user_id__icontains=search) |
                Q(status__icontains=search) |
                Q(payment_method__icontains=search)
            )

        queryset = queryset.filter(q_objects)
        
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
            
        if limit:
            try:
                limit = int(limit)
                queryset = queryset[:limit]
            except ValueError:
                pass

        if export:
            export_data = AdminPaymentSerializer(queryset, many=True).data
            if export == "csv":
                return self._export_csv_payment(export_data, 'payment_report')
            elif export == "pdf":
                return self._export_pdf_payment(export_data, 'payment_report')
            elif export == "xlsx":
                return self._export_xlsx_payment(export_data, 'payment_report')

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        
        serializer = AdminPaymentSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def _export_csv_payment(self, data, filename_prefix):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        
        writer.writerow(['ID', 'From User', 'Username', 'Linked Username', 'Level', 'Amount', 'GIC', 'Status', 'Payment Method', 'Created At'])
        
        for item in data:
            writer.writerow([
                item['id'], item['from_user'], item['username'], item['linked_username'], 
                item['level'], str(item['amount']), str(item['gic']), item['status'], 
                item['payment_method'], str(item['created_at'])
            ])
        return response

    def _export_pdf_payment(self, data, filename_prefix):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"{filename_prefix.replace('_', ' ').title()} Report", styles["Title"]))

        table_data = [['ID', 'From User', 'Username', 'Linked Username', 'Level', 'Amount', 'GIC', 'Status', 'Payment Method', 'Created At']]
        for item in data:
            table_data.append([
                str(item['id']), item['from_user'], item['username'], item['linked_username'], 
                item['level'], str(item['amount']), str(item['gic']), item['status'], 
                item['payment_method'], str(item['created_at'])
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _export_xlsx_payment(self, data, filename_prefix):
        wb = Workbook()
        ws = wb.active
        ws.title = "PaymentReport"
        ws.append(['ID', 'From User', 'Username', 'Linked Username', 'Level', 'Amount', 'GIC', 'Status', 'Payment Method', 'Created At'])
        for item in data:
            ws.append([
                item['id'], item['from_user'], item['username'], item['linked_username'], 
                item['level'], item['amount'], item['gic'], item['status'], 
                item['payment_method'], str(item['created_at'])
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

class AdminNotificationsView(ListAPIView):
    queryset = AdminNotification.objects.all().order_by('-timestamp')
    # Use the individual model serializer here
    serializer_class = AdminNotificationSerializer 
    permission_classes = [IsAdminUser]
    pagination_class = PageNumberPagination





class AdminAnalyticsView(APIView):
    permission_classes = [IsAdminUser]

    # --- 1. CORE QUERY UTILITY (Most Time Effective Improvement) ---
    def _get_user_statistics_queryset(self, request):
        """
        Builds the complex, annotated queryset once. This is the single source 
        of truth for all user statistics data (JSON, CSV, PDF, XLSX).
        """
        queryset = CustomUser.objects.all().order_by('date_of_joining')
        
        # Define Subqueries for related counts (efficiently handles one-to-many counts)
        referral_count_subquery = CustomUser.objects.filter(
            sponsor_id=OuterRef('user_id')
        ).values('sponsor_id').annotate(
            count=Count('id')
        ).values('count')[:1] 
        
        placement_count_subquery = CustomUser.objects.filter(
            placement_id=OuterRef('user_id')
        ).values('placement_id').annotate(
            count=Count('id')
        ).values('count')[:1]

        # Apply Annotations (The heavy lifting)
        queryset = queryset.annotate(
            # FIX: Explicit output_field=IntegerField()
            total_referrals=Subquery(referral_count_subquery, output_field=IntegerField()), 
            current_level_placement_count=Subquery(placement_count_subquery, output_field=IntegerField()),
            
            total_income_generated=Sum(
                'userlevel__received',
                # Ensure the output type is correct for decimal data
                output_field=DecimalField(max_digits=12, decimal_places=2) 
            ),

            levels_completed=Count(
                'userlevel',
                filter=Q(userlevel__status='paid') & ~Q(userlevel__level__name='Refer Help'),
                distinct=True
            ),
            
            total_payments_made=Sum(
                'userlevel__payments__amount',
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        ).prefetch_related('userlevel_set__level')

        # Apply Filters (must be applied to the annotated queryset)
        user_id_search = request.query_params.get('user_id')
        if user_id_search:
            queryset = queryset.filter(user_id__icontains=user_id_search)
        search_query = request.query_params.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(user_id__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query)
            )
        levels_completed_filter = request.query_params.get('levels_completed')
        if levels_completed_filter:
            try:
                # Filter against the 'levels_completed' annotation
                queryset = queryset.filter(levels_completed=int(levels_completed_filter))
            except ValueError:
                # Safely ignore invalid filter for exports, but could raise error for JSON/default
                pass 
                
        return queryset

    # --- 2. MAIN ENTRY POINT ---
    def get(self, request):
        report_type = request.query_params.get('report', 'summary').lower()
        export = request.query_params.get("export") # Check for export parameter

        try:
            if report_type == 'summary':
                return self._get_summary_analytics(request)
            
            elif report_type == 'user_stats':
                # 💥 Fetch the annotated queryset once, applying all filters
                queryset = self._get_user_statistics_queryset(request)
                
                if export:
                    # Route to the export handler (no pagination)
                    return self._handle_user_stats_export(queryset, export)
                else:
                    # Route to the standard JSON handler (with pagination)
                    # Renamed from _get_user_statistics to clarify its purpose
                    return self._get_user_statistics_paginated(request, queryset)
            
            return Response({"detail": "Invalid report type specified. Use 'summary' or 'user_stats'."}, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            # 💥 DEBUGGING: Print the exact exception for your logs 
            import traceback # Import traceback at the top of your file if needed
            print(f"--- CRITICAL GLOBAL ERROR in AdminAnalyticsView GET ---")
            print(f"Error Type: {type(e).__name__}")
            print(f"Error Details: {e}")
            # Optional: print(traceback.format_exc())
            print(f"-------------------------------------------------------")

            return Response({"detail": "An unexpected server error occurred during report generation."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    # --- 3. SUMMARY ANALYTICS (Corrected/Optimized) ---
    def _get_summary_analytics(self, request):
        # This function is already highly optimized using database aggregates, 
        # so no structural change is needed here.
        try:
            total_registered_users = CustomUser.objects.count()
            total_active_users = CustomUser.objects.filter(is_active=True).count()
            
            notification_stats = AdminNotification.objects.filter(
                operation_type='level_payment'
            ).aggregate(
                total_revenue_notified=Sum('amount'),
                total_gic_collected=Sum('gic')
            )

            total_revenue_paid = notification_stats['total_revenue_notified'] or Decimal('0.00')
            total_gic_collected = notification_stats['total_gic_collected'] or Decimal('0.00')

            income_stats = UserLevel.objects.filter(
                status='paid'
            ).aggregate(
                total_income_received_sum=Sum('received')
            )
            total_received_income = income_stats['total_income_received_sum'] or Decimal('0.00')
            
            users_per_level = UserLevel.objects.filter(
                status='paid', level__isnull=False
            ).values('level__name', 'level__order').annotate(
                count=Count('user_id', distinct=True)  
            ).order_by('level__order')

            users_by_level_dict = {
                item['level__name']: item['count'] 
                for item in users_per_level if item.get('level__name')
            }
            
            data = {
                'total_registered_users': total_registered_users,
                'total_active_users': total_active_users,
                'total_revenue_paid': total_revenue_paid.quantize(Decimal('0.01')),
                'total_gic_collected': total_gic_collected.quantize(Decimal('0.01')),
                'total_received_income': total_received_income.quantize(Decimal('0.01')),
                'Completed_users_by_level': users_by_level_dict,
            }
            
            serializer = AdminSummaryAnalyticsSerializer(data=data) 
            serializer.is_valid(raise_exception=True) 
            return Response(serializer.data)

        except (FieldError, OperationalError, ProgrammingError) as e:
            error_msg = "Database/Model configuration error while calculating summary analytics. "
            print(f"{error_msg} Details: {e}")
            return Response({"detail": error_msg, "error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            print(f"Unexpected error in _get_summary_analytics: {e}")
            return Response({"detail": "An unexpected error occurred in summary calculation."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # --- 4. USER STATISTICS (PAGINATED JSON) ---
    def _get_user_statistics_paginated(self, request, queryset):
        """ Handles standard JSON response and pagination using the pre-queried queryset. """
        try:
            # 1. Pagination setup
            paginator = PageNumberPagination()
            limit_param = request.query_params.get('limit')
            
            if limit_param:
                try:
                    paginator.page_size = int(limit_param)
                except ValueError:
                    return Response({"detail": "Invalid limit parameter. Must be an integer."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                paginator.page_size = 10
            
            page = paginator.paginate_queryset(queryset, request)
            
            # 2. Data formatting for JSON response
            data = []
            for user in page:
                # The annotations (total_income_generated, levels_completed, etc.) are available directly on the 'user' object now
                full_name_candidate = f"{user.first_name} {user.last_name}".strip()
                final_full_name = full_name_candidate or user.first_name or user.user_id 
                
                data.append({
                    'user_id': user.user_id,
                    'full_name': final_full_name,
                    'total_income_generated': user.total_income_generated or Decimal('0.00'),
                    'total_referrals': user.total_referrals or 0,
                    'current_level': user.current_level_placement_count or 0, # Uses annotation
                    'levels_completed': user.levels_completed or 0,
                    'total_payments_made': user.total_payments_made or Decimal('0.00'),
                })

            # FIX: Explicitly using 'data=' keyword remains correct here.
            serializer = UserAnalyticsSerializer(data=data, many=True)
            serializer.is_valid(raise_exception=True) 
            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            error = (f"Unexpected error in _get_user_statistics_paginated: {e}")
            return Response({"detail": error}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _get_user_stats_export_data(self, queryset):
  
        data = []
        
        for user in queryset.iterator(chunk_size=2000): 
            
            full_name_candidate = f"{user.first_name} {user.last_name}".strip()
            final_full_name = full_name_candidate or user.first_name or user.user_id 
            
            data.append({
                'user_id': user.user_id,
                'full_name': final_full_name,
                'email': user.email,
                'is_active': user.is_active,
                'total_income_generated': user.total_income_generated or Decimal('0.00'),
                'total_referrals': user.total_referrals or 0,
                'current_level_placement_count': user.current_level_placement_count or 0,
                'levels_completed': user.levels_completed or 0,
                'total_payments_made': user.total_payments_made or Decimal('0.00'),
            })
        return data


    # --- 5. EXPORT ROUTER (Receives pre-queried queryset) ---
    def _handle_user_stats_export(self, queryset, export_format):
        """Routes the queryset to the correct file format handler."""
        if export_format in ["csv", "xlsx"]:
            data = self._get_user_stats_export_data(queryset)
            
            if export_format == "csv":
                return self._export_csv_user_stats(data, 'user_analytics_report')
            elif export_format == "xlsx":
                return self._export_xlsx_user_stats(data, 'user_analytics_report')
        elif export_format == "pdf":
            return self._export_pdf_user_stats(queryset, 'user_analytics_report')

        
        return Response({"detail": "Invalid export format specified."}, status=status.HTTP_400_BAD_REQUEST)




    def _export_csv_user_stats(self, data, filename_prefix):
        """ Exports a pre-formatted list of dictionaries ('data') to CSV. """
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        
        # Headers must match the keys generated in the calling function (_handle_user_stats_export_data)
        headers = [
            'User ID', 'Full Name', 'Email', 'Active', 'Joined Date', 
            'Income Generated', 'Total Referrals', 'Placement Count', 
            'Levels Completed', 'Payments Made'
        ]
        writer.writerow(headers)
        
        for item in data:
            writer.writerow([
                item['user_id'], 
                item['full_name'], 
                item['email'], 
                item['is_active'],  
                str(item['total_income_generated']), # Convert Decimal/other types to string
                item['total_referrals'], 
                item['current_level_placement_count'], 
                item['levels_completed'], 
                str(item['total_payments_made'])
            ])
        return response

    def _export_xlsx_user_stats(self, data, filename_prefix):
        """ Exports a pre-formatted list of dictionaries ('data') to XLSX. """
        from openpyxl import Workbook # Assuming this is imported
        
        wb = Workbook()
        ws = wb.active
        ws.title = "UserAnalyticsReport"
        
        # Headers must match the keys generated in the calling function
        headers = [
            'User ID', 'Full Name', 'Email', 'Active', 'Joined Date', 
            'Income Generated', 'Total Referrals', 'Placement Count', 
            'Levels Completed', 'Payments Made'
        ]
        ws.append(headers)
        
        for item in data:
            ws.append([
                item['user_id'], 
                item['full_name'], 
                item['email'], 
                item['is_active'], 
                item['total_income_generated'], # Openpyxl handles Decimal/float better than csv
                item['total_referrals'], 
                item['current_level_placement_count'], 
                item['levels_completed'], 
                item['total_payments_made']
            ])
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    # NOTE: The _export_pdf_user_stats method already uses an in-memory list (table_data), 
    # but it was generated inside the function. We will update the router to handle it too.

    def _export_pdf_user_stats(self, queryset, filename_prefix):
        """ Uses ReportLab. Now includes placement count and payments made. """
        buffer = io.BytesIO()
        # NOTE: Ensure A4 is imported from reportlab.lib.pagesizes
        doc = SimpleDocTemplate(buffer, pagesize=A4) 
        elements = []
        # NOTE: Ensure getSampleStyleSheet is imported from reportlab.lib.styles
        styles = getSampleStyleSheet() 
        elements.append(Paragraph(f"{filename_prefix.replace('_', ' ').title()}", styles["Title"]))

        # FIX: Added Placement Count and Payments Made headers
        table_data = [
            ['User ID', 'Full Name', 'Income', 'Referrals', 'Placement Count', 'Levels Comp.', 'Total Paid']
        ]
        
        # FIX: Added 'current_level_placement_count' and 'total_payments_made' to .values()
        pdf_fields = [
            'user_id', 'first_name', 'last_name', 'total_income_generated', 
            'total_referrals', 'levels_completed', 
            'current_level_placement_count', 'total_payments_made'
        ]
        
        for user in queryset.values(*pdf_fields):
            full_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
            table_data.append([
                user.get('user_id', ''),
                full_name,
                str(user.get('total_income_generated', Decimal('0.00'))),
                str(user.get('total_referrals', 0)),
                str(user.get('current_level_placement_count', 0)), # FIX: Added Placement Count
                str(user.get('levels_completed', 0)),
                str(user.get('total_payments_made', Decimal('0.00'))), # FIX: Added Total Paid
            ])
            
        table = Table(table_data)
        # NOTE: Ensure colors is imported from reportlab.lib
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response