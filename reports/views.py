from rest_framework import viewsets, status
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework.decorators import action
from django.http import HttpResponse
import csv
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from django_filters.rest_framework import DjangoFilterBackend
from level.models import UserLevel, LevelPayment
from level.serializers import AdminPaymentReportSerializer
from .filters import PaymentFilter
from .serializers import DashboardReportSerializer, LevelPaymentReportSerializer, SendRequestReportSerializer, AUCReportSerializer, PaymentReportSerializer, LevelUsersSerializer,UserBonusListSerializer,BonusSummaryDataSerializer
from users.models import CustomUser
from django.db.models import Count, Sum, Q,Value, CharField
from django.utils import timezone
from datetime import timedelta
from rest_framework.views import APIView
from openpyxl import Workbook
from rest_framework.pagination import PageNumberPagination
import logging
from datetime import datetime
from django.db.models.functions import Concat
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticated
from decimal import Decimal
from django.conf import settings

logger = logging.getLogger(__name__)

# Existing PaymentReportViewSet (unchanged for now)
class PaymentReportPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class PaymentReportViewSet(viewsets.ReadOnlyModelViewSet):
    # N+1 FIX: Use select_related for speed
    queryset = UserLevel.objects.all().select_related('user', 'level').order_by('-requested_date') 
    serializer_class = AdminPaymentReportSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [DjangoFilterBackend]
    filterset_class = PaymentFilter
    
    pagination_class = PaymentReportPagination 


    @action(detail=False, methods=['get'], url_path='pending')
    def pending_payments(self, request):
        """
        List all strictly pending payments (status='pending' on UserLevel).
        """
        # ONLY filter by status='pending' on the UserLevel
        queryset = self.filter_queryset(self.get_queryset().filter(status='pending'))
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'], url_path='approved')
    def approved_payments(self, request):
        """
        List all approved (paid) payments with pagination and filtering.
        Endpoint: /api/payment-report/approved/
        """
        queryset = self.filter_queryset(self.get_queryset().filter(status='paid'))
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def list(self, request, *args, **kwargs):
        # Start with the optimized base queryset and apply request filters
        filtered_qs = self.filter_queryset(self.get_queryset())
        
        paginator = self.paginator
        
        # Store original param to ensure it's restored after BOTH lists are processed
        original_page_query_param = paginator.page_query_param
        
        # --- A. Paginate Pending List (uses ?p_page=X) ---
        paginator.page_query_param = 'p_page' # 🌟 NEW: Set param for PENDING
        
        pending_qs = filtered_qs.filter(status='pending') 
        page_pending = paginator.paginate_queryset(pending_qs, request=request, view=self)
        serializer_pending = self.get_serializer(page_pending, many=True)
        
        pending_data = {
            'count': paginator.page.paginator.count if paginator.page else 0,
            'next': paginator.get_next_link(), # This link now uses the 'p_page' parameter
            'previous': paginator.get_previous_link(),
            'results': serializer_pending.data,
        }

        # --- B. Paginate Approved List (uses ?a_page=X) ---
        
        # Set param for APPROVED
        paginator.page_query_param = 'a_page' 

        approved_qs = filtered_qs.filter(status='paid') 
        page_approved = paginator.paginate_queryset(approved_qs, request=request, view=self)
        serializer_approved = self.get_serializer(page_approved, many=True)
        
        approved_data = {
            'count': paginator.page.paginator.count if paginator.page else 0,
            'next': paginator.get_next_link(), # This link now uses the 'a_page' parameter
            'previous': paginator.get_previous_link(),
            'results': serializer_approved.data,
        }
        
        paginator.page_query_param = original_page_query_param

        return Response({
            'pending': pending_data,
            'approved': approved_data,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="payment_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Username', 'Level', 'Amount', 'Payment_Mode', 'Transaction_ID', 'Status', 'requested_date'])
        for obj in queryset:
            writer.writerow([
                getattr(obj.user, 'email', getattr(obj.user, 'user_id', 'Unknown')),
                obj.level.name,
                obj.level.amount,
                obj.payment_mode,
                obj.transaction_id or '',
                obj.status,
                obj.requested_date.strftime('%Y-%m-%d %H:%M:%S') if obj.requested_date else ''
            ])
        return response

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)
        y = 750
        p.drawString(100, y, "Payment Report")
        y -= 20
        p.drawString(100, y, f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} IST")
        y -= 40
        headers = ['Username', 'Level', 'Amount', 'Payment_Mode', 'Transaction_ID', 'Status', 'requested_date']
        for i, header in enumerate(headers):
            p.drawString(100 + i * 80, y, header)
        y -= 20
        for obj in queryset:
            row = [
                getattr(obj.user, 'email', getattr(obj.user, 'user_id', 'Unknown')),
                obj.level.name,
                str(obj.level.amount),
                obj.payment_mode,
                obj.transaction_id or '',
                obj.status,
                obj.requested_date.strftime('%Y-%m-%d %H:%M:%S') if obj.requested_date else ''
            ]
            for i, value in enumerate(row):
                p.drawString(100 + i * 80, y, value)
            y -= 20
            if y < 50:
                p.showPage()
                y = 750
        p.showPage()
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="payment_report.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='export-payment-details-csv')
    def export_payment_details_csv(self, request):
        queryset = LevelPayment.objects.all().order_by('-created_at')
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="payment_details_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Payment_ID', 'Payment_Token', 'Level_Name', 'User_ID', 'Username', 'Amount', 'Status', 
                        'Razorpay_Order_ID', 'Razorpay_Payment_ID', 'Razorpay_Signature', 'Created_At'])
        for obj in queryset:
            writer.writerow([
                obj.id,
                str(obj.payment_token),
                obj.user_level.level.name,
                obj.user_level.user.user_id,
                getattr(obj.user_level.user, 'email', getattr(obj.user_level.user, 'user_id', 'Unknown')),
                obj.amount,
                obj.status,
                obj.razorpay_order_id or '',
                obj.razorpay_payment_id or '',
                obj.razorpay_signature or '',
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else ''
            ])
        return response

    @action(detail=False, methods=['get'], url_path='export-payment-details-pdf')
    def export_payment_details_pdf(self, request):
        queryset = LevelPayment.objects.all().order_by('-created_at')
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)
        y = 750
        p.drawString(100, y, "Payment Details Report")
        y -= 20
        p.drawString(100, y, f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} IST")
        y -= 40
        headers = ['Payment_ID', 'Payment_Token', 'Level_Name', 'User_ID', 'Username', 'Amount', 'Status', 
                  'Razorpay_Order_ID', 'Razorpay_Payment_ID', 'Razorpay_Signature', 'Created_At']
        for i, header in enumerate(headers):
            p.drawString(100 + i * 60, y, header)
        y -= 20
        for obj in queryset:
            row = [
                str(obj.id),
                str(obj.payment_token),
                obj.user_level.level.name,
                obj.user_level.user.user_id,
                getattr(obj.user_level.user, 'email', getattr(obj.user_level.user, 'user_id', 'Unknown')),
                str(obj.amount),
                obj.status,
                obj.razorpay_order_id or '',
                obj.razorpay_payment_id or '',
                obj.razorpay_signature or '',
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else ''
            ]
            for i, value in enumerate(row):
                p.drawString(100 + i * 60, y, value)
            y -= 20
            if y < 50:
                p.showPage()
                y = 750
        p.showPage()
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="payment_details_report.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='custom-user-report')
    def custom_user_report(self, request):
        logger.debug("Custom user report endpoint hit")
        print("Request path:", request.path)  # Debug the incoming path
        users = CustomUser.objects.filter(is_active=True).prefetch_related('userlevel_set')
        report_data = []

        for user in users:
            user_levels = UserLevel.objects.filter(user=user)
            completed_levels = user_levels.filter(status='paid', level__order__lte=6).count()
            total_received = user_levels.aggregate(total=Sum('received'))['total'] or 0
            received_count = user_levels.filter(status='paid').count()
            refer_help = user_levels.filter(level__name='Refer Help').first()
            refer_help_amount = refer_help.level.amount if refer_help else 0
            refer_help_paid_amount = refer_help_amount if refer_help and refer_help.status == 'paid' else 0
            total_amount_sent = user_levels.filter(status='paid').aggregate(total=Sum('level__amount'))['total'] or 0
            total_amount_received = total_received

            report_data.append({
                'username': f"{user.first_name} {user.last_name}".strip() or user.email,
                'level_completed': completed_levels,
                'total_received': total_received,
                'received_count': received_count,
                'refer_help_amount': refer_help_amount,
                'amount_if_paid': refer_help_paid_amount,
                'total_amount_sent': total_amount_sent,
                'total_amount_received': total_amount_received
            })

        return Response(report_data)

# Existing DashboardReportViewSet (unchanged)
# Assuming necessary imports (Decimal, Sum, Count, Q, timezone, timedelta) are present

class DashboardReportViewSet(viewsets.ViewSet):
    permission_classes = [IsAdminUser]

    def list(self, request):
        
        # Define constants
        PAID_STATUS = 'paid'
        VERIFIED_STATUS = 'Verified'

        # Helper function MUST be defined here, or imported from utils, 
        # but CANNOT contain the ORM queries inside it.
        def safe_user_data(user_level_instance):
            """Safely extracts user data from a UserLevel instance or returns N/A dict."""
            if not user_level_instance or not user_level_instance.user:
                return {'name': 'N/A', 'email_id': 'N/A', 'first_name': 'N/A', 'last_name': 'N/A', 'amount': 0, 'time': 'N/A'}

            user = user_level_instance.user
            level = user_level_instance.level

            # NOTE: Removed the redundant ORM queries that were causing the flaw.
            return {
                'name': f"{user.first_name} {user.last_name}".strip() or user.email,
                'email_id': user.email or 'N/A',
                'first_name': user.first_name or 'N/A',
                'last_name': user.last_name or 'N/A',
                'amount': level.amount or 0,
                'time': user_level_instance.requested_date.strftime('%Y-%m-%d %H:%M:%S') if user_level_instance.requested_date else 'N/A'
            }
        
        # 1. Total Members... (Correct)
        total_members = CustomUser.objects.filter(is_active=True).count()
        logger.debug(f"Total members: {total_members}")
        
        # 2. Total Income... (Correct: Filters by PAID_STATUS)
        total_income = UserLevel.objects.filter(
            status=PAID_STATUS  
        ).aggregate(total=Sum('received'))['total']
        total_income = total_income if isinstance(total_income, (int, float, Decimal)) else Decimal(0)
        logger.debug(f"Total income: {total_income}")

        # 3. Total Active Level 6... (Correct)
        active_level_6 = CustomUser.objects.filter(
            userlevel__level__order__lte=6,
            userlevel__status=PAID_STATUS
        ).annotate(
            paid_levels=Count('userlevel', filter=Q(userlevel__level__order__lte=6, userlevel__status=PAID_STATUS))
        ).filter(paid_levels=6).distinct().count()
        logger.debug(f"Total active level 6: {active_level_6}")

        # 4. New users on each level... (Correct)
        new_users_per_level = []
        for lvl in range(1, 7):
            users_at_level = CustomUser.objects.filter(
                userlevel__level__order__lte=lvl,
                userlevel__status=PAID_STATUS
            ).annotate(
                paid_levels=Count('userlevel', filter=Q(userlevel__level__order__lte=lvl, userlevel__status=PAID_STATUS))
            ).filter(paid_levels=lvl).distinct().count()
            new_users_per_level.append({'level': lvl, 'count': users_at_level})
            logger.debug(f"New users at level {lvl}: {users_at_level}")

        # 5. Recent payments... (Correct: select_related and status filter are correct)
        recent_payments_qs = LevelPayment.objects.filter(
            status=VERIFIED_STATUS,
            created_at__gte=timezone.now() - timedelta(days=30)
        ).select_related('user_level__user').order_by('-created_at')[:10]
        # Replaced duplicated serializer line with the correct assignment:
        recent_payments = LevelPaymentReportSerializer(recent_payments_qs, many=True).data
        logger.debug(f"Recent payments count: {len(recent_payments)}")

        # 6. New user registrations... (Correct)
        recent_users_qs = CustomUser.objects.filter(
            date_of_joining__gte=timezone.now() - timedelta(days=30)
        ).order_by('-date_of_joining').prefetch_related('userlevel_set')[:10] 
        
        new_user_registrations = []
        for user in recent_users_qs:
            completed_levels = UserLevel.objects.filter(
                user=user, status=PAID_STATUS, level__order__lte=6
            ).count()
            
            username_display = f"{user.first_name} {user.last_name}".strip() 
            new_user_registrations.append({
                'user_id': user.user_id,
                'username': username_display or user.email, 
                'levels_done': completed_levels
            })
        logger.debug(f"New user registrations count: {len(new_user_registrations)}")
        
        # 7. Latest report: Data for latest help requests and payment
        # Queries are now correct and executed ONCE
        latest_refer_help = UserLevel.objects.filter(
            level__name='Refer Help',  
            status=PAID_STATUS 
        ).order_by('-requested_date').select_related('user', 'level').first()

        latest_level_help = UserLevel.objects.filter(
            level__name__contains='Level',  
            status=PAID_STATUS 
        ).order_by('-requested_date').select_related('user', 'level').first()

        latest_level_payment = LevelPayment.objects.filter(
            status=VERIFIED_STATUS # ✅ FIX: Only verified payments are considered the "latest"
        ).order_by('-created_at').select_related('user_level__user').first()

        latest_report = {
            'latest_refer_help': latest_refer_help.level.name if latest_refer_help and latest_refer_help.level else 'N/A',
            
            'latest_refer_user': {
                **safe_user_data(latest_refer_help),
                'status': latest_refer_help.status if latest_refer_help else 'N/A'
            },
            
            'latest_level_help': latest_level_help.level.name if latest_level_help and latest_level_help.level else 'N/A',
            
            'latest_level_payment': {
                'amount': latest_level_payment.amount if latest_level_payment else 0,
                'time': latest_level_payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if latest_level_payment and latest_level_payment.created_at else 'N/A',
                'status': latest_level_payment.status if latest_level_payment else 'N/A',
                'done': latest_level_payment.status == VERIFIED_STATUS if latest_level_payment else False
            }
        }
        
        # 8. Final Response Serialization... (Correct)
        data = {
            'total_members': total_members,
            'total_income': total_income,
            'total_active_level_6': active_level_6,
            'new_users_per_level': new_users_per_level,
            'recent_payments': recent_payments,
            'new_user_registrations': new_user_registrations,
            'latest_report': latest_report
        }
        
        serializer = DashboardReportSerializer(data) 
        return Response(serializer.data)

# Existing UserReportViewSet (unchanged)
class UserReportViewSet(viewsets.ViewSet):
    # Recommended: Enforce authentication
    permission_classes = [IsAuthenticated] 

    @action(detail=False, methods=['get'], url_path='user-report')
    def user_report(self, request):
        logger.debug("User report endpoint hit for user %s", request.user.user_id)
        user = request.user
        
        # Use reverse relation, caching the queryset for reuse
        # Assuming the reverse relation name from CustomUser to UserLevel is 'userlevel_set'
        user_levels = user.userlevel_set.all() 
        
        # --- Calculations ---
        
        # 1. Total Income (Total Received) - Sum of 'received' across ALL UserLevel records.
        # This is the most accurate definition of total income.
        # Use .aggregate() only once for efficiency if possible, or directly on the queryset
        current_user_id_str = str(user.user_id)
        total_referral_income = UserLevel.objects.filter(
                # Filter 1: Match the recipient ID (uses the CharField)
                linked_user_id=current_user_id_str, 
                level__name='Refer Help',
                status='paid' 
            ).aggregate(
                total=Sum('level__amount')
            )['total'] or Decimal(0)
        
        # 2. Total Paid (Send Help) - Sum of 'level__amount' for completed (paid) matrix levels (1-6)
        total_paid_for_levels = user_levels.filter(
            status='paid',
            level__order__lte=6 
        ).aggregate(
            total=Sum('level__amount')
        )['total'] or Decimal(0)
        


        # 3. Levels Completed - Count of paid matrix levels (1-6)
        completed_levels = user_levels.filter(
            status='paid', 
            level__order__lte=6
        ).count()

        received_help_amount = user_levels.filter().exclude(
                level__name='Refer Help'
            ).aggregate(
                    total=Sum('received')
            )['total'] or Decimal(0)
                
        # 4. Pending Counts - Count of levels that are currently pending payment/approval.
        pending_count = user_levels.filter(status='pending').count()
        total_income = received_help_amount + total_referral_income

        total_amount_generated= received_help_amount  - total_paid_for_levels
        
        # 5. Referral Count - FIXING THE AttributeError
        # The correct, explicit query is needed here.
        try:
            # Assuming CustomUser is the model where 'sponsor_id' links to 'user_id'
            referral_count = CustomUser.objects.filter(sponsor_id=user.user_id).count() 
        except Exception as e:
            logger.error(f"Error counting referrals for user {user.user_id}: {e}")
            referral_count = 0

        # --- Data Consolidation ---
        
        data = {
            'username': f"{user.first_name} {user.last_name}".strip() or user.email,
            'level_completed': completed_levels,
            
            # Income metrics (Total Received / Total Income / Total Amount Generated are all the same)
            'total_received': total_amount_generated, 
            'total_amount_generated': total_income ,
            "total_referral_income":total_referral_income, 
            
            # Pending metrics
            'pending_send_count': pending_count,
            'pending_receive_count': pending_count, # Assuming pending status covers both send/receive for simplicity
            
            # Payment metrics
            'send_help': total_paid_for_levels, # Total amount paid by the user
            'receive_help': received_help_amount,   
            'referral_count': referral_count
        }
        return Response(data)

    # ----------------------------------------------------------------------

    @action(detail=False, methods=['get'], url_path='total-payment-info')
    def total_payment_info(self, request):
        user = request.user
        # Use reverse relation
        user_levels = user.userlevel_set.all() 
        
        # Total Amount Received (Total Income)
        total_amount_received = user_levels.aggregate(total=Sum('received'))['total'] or Decimal(0)
        
        # Total Payments Made by the User (Send Help)
        total_payments_made = user_levels.filter(
            status='paid',
            level__order__lte=6 
        ).aggregate(
            total=Sum('level__amount')
        )['total'] or Decimal(0)
        
        # Balance Left (Net Profit/Loss)
        balance_left = total_amount_received - total_payments_made
        
        data = {
            'total_amount_received': total_amount_received,
            'total_payments_made': total_payments_made,
            'balance_left': balance_left
        }
        return Response(data)

    @action(detail=False, methods=['get'], url_path='downline-level-count')
    def downline_level_count(self, request):
        user = request.user
        ADMIN_SPONSOR_ID = getattr(settings, 'ADMIN_USER_ID', 'ADMIN001')

        # 1. --- EFFICIENTLY GET ALL DOWNLINE USER IDs (Iterative Bulk Query) ---
        
        # Use a set for O(1) lookups and to guarantee unique IDs
        downline_user_ids = set()
        
        # Start the list of IDs to process with the current user's direct referrals
        current_level_ids = list(
            CustomUser.objects.filter(sponsor_id=user.user_id)
            .exclude(sponsor_id=ADMIN_SPONSOR_ID)
            .values_list('user_id', flat=True)
        )
        
        processed_ids = {user.user_id, ADMIN_SPONSOR_ID}

        # Loop until no new referrals are found in the next level
        while current_level_ids:
            # Add the IDs found in the current level to the overall set
            new_ids = set(current_level_ids) - processed_ids
            
            # Stop processing if no genuinely new users were found
            if not new_ids:
                break

            downline_user_ids.update(new_ids)
            processed_ids.update(new_ids)
            
            # Get the next level of referrals in ONE bulk query
            current_level_ids = list(
                CustomUser.objects.filter(
                    sponsor_id__in=new_ids
                ).exclude(sponsor_id=ADMIN_SPONSOR_ID)
                .values_list('user_id', flat=True)
            )

        # Convert the set back to a list for compatibility with __in query
        downline_user_ids_list = list(downline_user_ids)
        total_downline_members = len(downline_user_ids_list)

        # Handle case where downline is empty
        if total_downline_members == 0:
            empty_data = [{'level': i, 'count': 0, 'percentage': 0.00} for i in range(1, 7)]
            return Response({
                'paid_level_entries_by_level': empty_data,
                'total_downline_members': 0
            })

        # 2. --- COUNT TOTAL PAID LEVEL ENTRIES FOR DOWNLINE USERS (UNCHANGED) ---
        paid_level_counts_qs = UserLevel.objects.filter(
            user__user_id__in=downline_user_ids_list, # Use the computed list
            status='paid',
            level__order__lte=6
        ).values('level__order').annotate(
            total_paid_entries=Count('id')
        ).order_by('level__order')

        # 3. --- CALCULATE PERCENTAGE AND FORMAT OUTPUT (UNCHANGED) ---
        
        paid_level_entries_by_level = {i: 0 for i in range(1, 7)}
        
        for item in paid_level_counts_qs:
            level_num = item['level__order']
            if 1 <= level_num <= 6:
                paid_level_entries_by_level[level_num] = item['total_paid_entries']
                
        paid_level_entries_by_level_list = []
        for lvl, count in paid_level_entries_by_level.items():
            percentage = (count / total_downline_members) * 100.0 if total_downline_members else 0.00
            
            paid_level_entries_by_level_list.append({
                'level': lvl, 
                'count': count,
                'percentage': round(percentage, 2)
            })

        return Response({
            'paid_level_entries_by_level': paid_level_entries_by_level_list,
            'total_downline_members': total_downline_members,
        }, status=status.HTTP_200_OK)

class UserLatestReportView(APIView):
    # Pagination is defined but not used here since we are only fetching 'first()' elements.
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10

    def get(self, request, *args, **kwargs):
        # logger.debug("User latest report endpoint hit for user %s", request.user.user_id)
        if not request.user.is_authenticated:
            return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)

        user = request.user
        user_id_str = user.user_id
        
        # 1. EFFICIENTLY GET ALL TARGET USER IDs (User + Direct Referrals)
        # Use Q object to fetch all IDs in a single query.
        target_user_ids = list(
            CustomUser.objects.filter(
                Q(user_id=user_id_str) | Q(sponsor_id=user_id_str)
            ).values_list('user_id', flat=True)
        )
        
        # 2. OPTIMIZED LEVEL HELP QUERY (Refer Help + Latest Level Help)
        # Filter all relevant UserLevels in one go, pre-fetching related data.
        # This addresses multiple N+1 hazards in the final dictionary creation.
        
        user_levels_qs = UserLevel.objects.filter(
            user__user_id__in=target_user_ids
        ).select_related(
            'level',  # Pre-fetch Level details (name, amount)
            'user'    # Pre-fetch CustomUser details (first_name, last_name, email)
        ).order_by('-requested_date')
        
        # Find the latest requested date for all Level-based entries
        latest_level_help = user_levels_qs.filter(
            level__name__contains='Level'
        ).first()

        # Find the latest Refer Help entry
        latest_refer_help = user_levels_qs.filter(
            level__name='Refer Help'
        ).first()
        
        # 3. OPTIMIZED LEVEL PAYMENT QUERY
        # Fetch the latest LevelPayment record.
        latest_level_payment = LevelPayment.objects.filter(
            user_level__user__user_id__in=target_user_ids
        ).order_by('-created_at').first()
        
        # --- Data Construction (Unchanged Logic, but now fast due to pre-fetching) ---

        # Helper function to safely format date
        def format_date(date_obj):
            return date_obj.strftime('%Y-%m-%d %H:%M:%S') if date_obj else 'N/A'
        
        # Helper function to safely get user data
        def get_user_data(ul_instance, default_amount, default_level):
            if ul_instance:
                user_instance = ul_instance.user
                level_instance = ul_instance.level
                return {
                    'name': f"{user_instance.first_name or ''} {user_instance.last_name or ''}".strip(),
                    'email_id': user_instance.email or 'admin@gmail.com',
                    'first_name': user_instance.first_name or '',
                    'last_name': user_instance.last_name or '',
                    'amount': level_instance.amount,
                    'time': format_date(ul_instance.requested_date),
                    'user_id': user_instance.user_id
                }
            return {
                'name': '', 'email_id': 'admin@gmail.com', 'first_name': '', 'last_name': '', 
                'amount': default_amount, 'time': 'N/A', 'user_id': 'N/A'
            }

        latest_report = {
            'latest_refer_help': latest_refer_help.level.name if latest_refer_help else 'N/A',
            'latest_refer_user': get_user_data(latest_refer_help, 1000.0, 'Refer Help'),
            
            'latest_level_help': latest_level_help.level.name if latest_level_help else 'Level 2',
            'latest_level_user': get_user_data(latest_level_help, 200.0, 'Level 2'),
            
            'latest_level_payment': {
                'amount': latest_level_payment.amount if latest_level_payment else 200.0,
                'time': format_date(latest_level_payment.created_at) if latest_level_payment else '2025-09-18 16:15:19'
            }
        }

        data = {'latest_report': latest_report}
        return Response(data, status=status.HTTP_200_OK)

# New Report Views
class SendRequestReport(APIView):
    # permission_classes = [IsAdminUser]
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10

    def get(self, request):
        queryset = UserLevel.objects.select_related('user', 'level').filter(user=request.user).order_by('-approved_at')
        
        # Query params
        email = request.query_params.get("email")
        status = request.query_params.get("status")
        user_id = request.query_params.get("user_id")
        username = request.query_params.get("username")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")  # 'csv', 'pdf', 'xlsx'

        # Filters
        if email:
            queryset = queryset.filter(user__email__icontains=email.lower())
        if username:
            # Annotate with a concatenated username (first_name + space + last_name)
            queryset = queryset.annotate(
                full_username=Concat('user__first_name', Value(' '), 'user__last_name', output_field=CharField())
            ).filter(
                Q(full_username__icontains=username.lower()) |
                Q(user__first_name__icontains=username.lower()) |
                Q(user__last_name__icontains=username.lower())
            )
        if status and status.lower() != "all":
            if status.lower() == "completed":
                queryset = queryset.filter(status='paid')  
            elif status.lower() == "pending":
                queryset = queryset.exclude(status='paid')
        if user_id:
            queryset = queryset.filter(user__user_id__icontains=user_id)
        if start_date:
            queryset = queryset.filter(requested_date__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(requested_date__date__lte=end_date)

        # Search filter
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__user_id__icontains=search) |
                Q(user__user_id__icontains=search) |
                Q(level__name__icontains=search) |
                Q(status__icontains=search)
            )

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__gte=start_date)
            except ValueError:
                logger.error(f"Invalid start_date format: {start_date}")
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__lte=end_date)
            except ValueError:
                logger.error(f"Invalid end_date format: {end_date}")

        # Limit
        if limit:
            try:
                limit = int(limit)
                queryset = queryset[:limit]
            except ValueError:
                pass

        # Export options
        if export == "csv":
            return self.export_csv(queryset, 'send_request_report')
        elif export == "pdf":
            return self.export_pdf(queryset, 'send_request_report')
        elif export == "xlsx":
            return self.export_xlsx(queryset, 'send_request_report')

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = SendRequestReportSerializer(page, many=True, context={'request': request})
            return paginator.get_paginated_response(serializer.data)
        serializer = SendRequestReportSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)

    def export_csv(self, queryset, filename_prefix):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        writer.writerow(['From User', 'Username', 'From Name', 'Amount', 'Status', 'Requested At'])
        serializer = SendRequestReportSerializer(queryset, many=True)
        for data in serializer.data:
            writer.writerow([
                data['from_user'],
                data['username'],
                data['from_name'],
                str(data['amount']),
                data['status'],
                data['requested_date'] if data['requested_date'] else ''
            ])
        return response

    def export_pdf(self, queryset, filename_prefix):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"{filename_prefix.replace('_', ' ').title()} Report", styles["Title"]))

        data = [['From User', 'Username', 'From Name', 'Amount', 'Status', 'Approved At']]
        serializer = SendRequestReportSerializer(queryset, many=True)
        for data_item in serializer.data:
            data.append([
                data_item['from_user'],
                data_item['username'],
                data_item['from_name'],
                str(data_item['amount']),
                data_item['status'],
                data_item['requested_date'] if data_item['requested_date'] else ''
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        return response

    def export_xlsx(self, queryset, filename_prefix):
        wb = Workbook()
        ws = wb.active
        ws.title = "SendRequestReport"
        ws.append(['From User', 'Username', 'From Name', 'Amount', 'Status', 'Approved At'])
        serializer = SendRequestReportSerializer(queryset, many=True)
        for data in serializer.data:
            ws.append([
                data['from_user'],
                data['username'],
                data['from_name'],
                data['amount'],
                data['status'],
                data['requested_date'] if data['requested_date'] else ''
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response





class AUCReport(APIView):
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10
    # permission_classes = [IsAdminUser]

    def get(self, request):
        queryset = UserLevel.objects.select_related('user', 'level').filter(user=request.user).order_by('-approved_at')
        
        # Query params

        email = request.query_params.get("email")
        status = request.query_params.get("status")
        user_id = request.query_params.get("user_id")
        username = request.query_params.get("username")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")  # 'csv', 'pdf', 'xlsx'

        # Filters
        if email:
            queryset = queryset.filter(user__email__icontains=email.lower())
        if username:
            # Annotate with a concatenated username (first_name + space + last_name)
            queryset = queryset.annotate(
                full_username=Concat('user__first_name', Value(' '), 'user__last_name', output_field=CharField())
            ).filter(
                Q(full_username__icontains=username.lower()) |
                Q(user__first_name__icontains=username.lower()) |
                Q(user__last_name__icontains=username.lower())
            )
        if status and status.lower() != "all":
            if status.lower() == "completed":
                queryset = queryset.filter(status='paid')  
            elif status.lower() == "pending":
                queryset = queryset.exclude(status='paid')
        if user_id:
            queryset = queryset.filter(user__user_id__icontains=user_id)
        if start_date:
            queryset = queryset.filter(requested_date__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(requested_date__date__lte=end_date)


        # Search filter
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__user_id__icontains=search) |
                Q(user__user_id__icontains=search) |
                Q(status__icontains=search)
            )

        
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__gte=start_date)
            except ValueError:
                logger.error(f"Invalid start_date format: {start_date}")
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__lte=end_date)
            except ValueError:
                logger.error(f"Invalid end_date format: {end_date}")

        # Pagination
        # Limit
        if limit:
            try:
                limit = int(limit)
                queryset = queryset[:limit]
            except ValueError:
                pass

        # Export options
        if export == "csv":
            return self.export_csv(queryset, 'auc_report')
        elif export == "pdf":
            return self.export_pdf(queryset, 'auc_report')
        if export == "xlsx":
            return self.export_xlsx(request, queryset, 'auc_report')

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = AUCReportSerializer(page, many=True, context={'request': request})
            return paginator.get_paginated_response(serializer.data)
        serializer = AUCReportSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)

    def export_csv(self, queryset, filename_prefix):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        writer.writerow(['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Status', 'Date'])
        serializer = AUCReportSerializer(queryset, many=True)
        for data in serializer.data:
            writer.writerow([
                data['from_user'],
                data['username'],
                data['from_name'],
                data['linked_username'],
                str(data['amount']),
                data['status'],
                data['date'] if data['date'] else ''
            ])
        return response

    def export_pdf(self, queryset, filename_prefix):
        buffer = io.BytesIO()
        try:
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            y = height - 50
            p.setFont("Helvetica-Bold", 14)
            p.drawString(150, y, f"{filename_prefix.replace('_', ' ').title()} Report")
            y -= 30
            p.setFont("Helvetica", 10)

            # Header
            header = "From User | Username | From Name | Linked Username | Amount | Status | Date"
            p.drawString(50, y, header)
            y -= 20

            # Data
            serializer = AUCReportSerializer(queryset, many=True)
            for data in serializer.data:
                line = f"{data['from_user']} | {data['username']} | {data['from_name']} | {data['linked_username']} | {data['amount']} | {data['status']} | {data['date'] if data['date'] else ''}"
                p.drawString(50, y, line)
                y -= 20
                if y < 50:
                    p.showPage()
                    y = height - 50

            p.save()
            pdf = buffer.getvalue()
            buffer.close()

            response = HttpResponse(pdf, content_type="application/pdf")
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")  # e.g., 20250920_0114
            response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timestamp}.pdf"'
            response['Content-Length'] = len(pdf)
            logger.info(f"PDF exported successfully: {len(queryset)} records processed")
        except Exception as e:
            logger.error(f"Error generating PDF: {e}", exc_info=True)
            return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
        finally:
            if 'buffer' in locals() and buffer:
                buffer.close()
                logger.debug("Buffer closed")

        return response

    def export_xlsx(self, request, queryset, filename_prefix):
        wb = Workbook()
        ws = wb.active
        ws.title = "AUCReport"
        ws.append(['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Status', 'Date'])
        serializer = AUCReportSerializer(queryset, many=True, context={'request': request})  # Pass request to context
        for data in serializer.data:
            ws.append([
                data['from_user'],
                data['username'],
                data['from_name'],
                data['linked_username'],
                data['amount'],
                data['status'],
                data['date'] if data['date'] else ''
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")  # Define timestamp here
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timestamp}.xlsx"'
        return response



class PaymentReport(APIView):
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10
    # permission_classes = [IsAdminUser]

    def get(self, request):
        queryset = UserLevel.objects.select_related('user', 'level').filter(user=request.user).order_by('-requested_date')
        
 # Query params
        email = request.query_params.get("email")
        status = request.query_params.get("status")
        user_id = request.query_params.get("user_id")
        username = request.query_params.get("username")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        export = request.query_params.get("export")  # 'csv', 'pdf', 'xlsx'

        # Filters
        if email:
            queryset = queryset.filter(user__email__icontains=email.lower())
        if username:
            # Annotate with a concatenated username (first_name + space + last_name)
            queryset = queryset.annotate(
                full_username=Concat('user__first_name', Value(' '), 'user__last_name', output_field=CharField())
            ).filter(
                Q(full_username__icontains=username.lower()) |
                Q(user__first_name__icontains=username.lower()) |
                Q(user__last_name__icontains=username.lower())
            )
        if status and status.lower() != "all":
            if status.lower() == "completed":
                queryset = queryset.filter(status='paid')  
            elif status.lower() == "pending":
                queryset = queryset.exclude(status='paid')
        if user_id:
            queryset = queryset.filter(user__user_id__iexact=user_id)
        if start_date:
            queryset = queryset.filter(requested_date__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(requested_date__date__lte=end_date)


        # Search filter
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__user_id__icontains=search) |
                Q(status__icontains=search)
            )

       

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__gte=start_date)
            except ValueError:
                logger.error(f"Invalid start_date format: {start_date}")
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__lte=end_date)
            except ValueError:
                logger.error(f"Invalid end_date format: {end_date}")

         # Limit
        if limit:
            try:
                limit = int(limit)
                queryset = queryset[:limit]
            except ValueError:
                pass

        # Export options
        if export == "csv":
            return self.export_csv(queryset, 'payment_report')
        elif export == "pdf":
            return self.export_pdf(queryset, 'payment_report')
        elif export == "xlsx":
            return self.export_xlsx(queryset, 'payment_report')

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = PaymentReportSerializer(page, many=True, context={'request': request})
            return paginator.get_paginated_response(serializer.data)
        serializer = PaymentReportSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)


    def export_csv(self, queryset, filename_prefix):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        writer.writerow(['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Payout Amount', 'Transaction Fee', 'Status', 'Approved At', 'Total'])
        serializer = PaymentReportSerializer(queryset, many=True)
        for data in serializer.data:
            writer.writerow([
                data['from_user'],
                data['username'],
                data['from_name'],
                data['linked_username'],
                str(data['amount']),
                str(data['payout_amount']),
                str(data['transaction_fee']),
                data['status'],
                data['requested_date'] if data['requested_date'] else '',
                str(data['total'])
            ])
        return response

    def export_pdf(self, queryset, filename_prefix):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"{filename_prefix.replace('_', ' ').title()} Report", styles["Title"]))

        data = [['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Payout Amount', 'Transaction Fee', 'Status', 'Approved At', 'Total']]
        serializer = PaymentReportSerializer(queryset, many=True)
        for data_item in serializer.data:
            data.append([
                data_item['from_user'],
                data_item['username'],
                data_item['from_name'],
                data_item['linked_username'],
                str(data_item['amount']),
                str(data_item['payout_amount']),
                str(data_item['transaction_fee']),
                data_item['status'],
                data_item['requested_date'] if data_item['requested_date'] else '',
                str(data_item['total'])
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        return response

    def export_xlsx(self, queryset, filename_prefix):
        wb = Workbook()
        ws = wb.active
        ws.title = "PaymentReport"
        ws.append(['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Payout Amount', 'Transaction Fee', 'Status', 'Approved At', 'Total'])
        serializer = PaymentReportSerializer(queryset, many=True)
        for data in serializer.data:
            ws.append([
                data['from_user'],
                data['username'],
                data['from_name'],
                data['linked_username'],
                data['amount'],
                data['payout_amount'],
                data['transaction_fee'],
                data['status'],
                data['requested_date'] if data['requested_date'] else '',
                data['total']
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
class StandardPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class AllUserBonusSummaryListView(ListAPIView):
    """Admin endpoint to list all users with pagination, search, and PDF download links."""
    queryset = CustomUser.objects.all().order_by('id')
    serializer_class = UserBonusListSerializer
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated, IsAdminUser] # Restrict to admins/staff

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(user_id__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
        return queryset

class SingleUserBonusSummaryView(APIView):
    """Handles PDF generation and JSON summary for a single user (by ID)."""
    permission_classes = [IsAuthenticated] 

    def create_income_statement_pdf(self, data, user):
        """Generates the PDF Income Statement using the data calculated by the serializer."""
        buffer = io.BytesIO()
        # Set margins to be a bit tighter
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom style for bolding titles
        bold_style = styles['Normal'].clone('BoldStyle')
        bold_style.fontName = 'Helvetica-Bold'
        
        # --- 1. Header/User Info ---
        user_info_table = Table([
            [Paragraph('<font size=20>WINNERS CLUBX</font>', styles['Title']), '', f"UserID: {data['user_id']}"],
            ['', '', f"UserName: {data['username']}"],
        ], colWidths=[2.5*inch, 2*inch, 2*inch])
        
        user_info_table.setStyle(TableStyle([
            ('ALIGN', (2,0), (2,1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('FONTNAME', (2,0), (2,1), 'Helvetica-Bold'),
        ]))
        elements.append(user_info_table)
        elements.append(Paragraph('Income Statement', styles["Heading1"]))
        elements.append(Spacer(1, 12))
        
        # --- 2. Contact/Statement Date Info ---
        # NOTE: You may need to adjust attribute names (e.g., 'mobile', 'upi_number') 
        # to match your CustomUser/Profile model structure.
        phone = getattr(user, 'mobile', 'xxxxxxxxxx')
        email = getattr(user, 'email', 'N/A')
        # Assuming profile attribute exists and has a 'upi_number' field
        upi_number = getattr(getattr(user, 'profile', None), 'upi_number', '9562763166') 
        
        contact_info_table = Table([
            [f"Phone: {phone[:2]}{'x' * (len(phone)-3)}", f"Statement Date: {data['statement_date']}"],
            [f"Email: {email}", ''],
            [f"UPI Number: {upi_number}", ''],
        ], colWidths=[3*inch, 3*inch])
        
        contact_info_table.setStyle(TableStyle([
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        elements.append(contact_info_table)
        elements.append(Spacer(1, 24))

        # --- 3. Income Table ---
        income_data = [
            [Paragraph('Income', bold_style), Paragraph('Amount', bold_style)],
            ['Referral Bonus', f"{data['referral_bonus']:.2f}"],
            ['Level Help', f"{data['level_help']:.2f}"],
            ['Send Help', f"{data['send_help']:.2f}"],
            [Paragraph('Net Amount', bold_style), f"{data['net_amount']:.2f}"],
            [Paragraph('Received Total', bold_style), f"{data['received_total']:.2f}"],
            [Paragraph("Please Enter a amount", styles['Normal']), ''] # Footer text from image
        ]

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCCC')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, 7), 'RIGHT'), # Align amounts to the right
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 5), (-1, 7), 'Helvetica-Bold'), # Total rows
        ])
        
        income_table = Table(income_data, colWidths=[4*inch, 1.5*inch])
        income_table.setStyle(table_style)
        elements.append(income_table)
        elements.append(Spacer(1, 24))

        # 4. Footer
        elements.append(Paragraph("For More Details Visit At:", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    # The 'get' method remains unchanged and is correct
    def get(self, request, user_id):
        # 1. Fetch User and Authorize
        try:
            target_user = CustomUser.objects.get(user_id=user_id)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
        
        # Non-admins can only view their own report
        if not request.user.is_staff and target_user != request.user:
             return Response({"error": "You can only access your own summary."}, status=status.HTTP_403_FORBIDDEN)

        # 2. Get Summary Data via Serializer
        # Make sure BonusSummaryDataSerializer is correctly imported
        serializer = BonusSummaryDataSerializer(target_user)
        summary_data = serializer.data
        
        # 3. Handle PDF Export
        export = request.query_params.get("export", "").lower() 
        if export == "pdf":
            try:
                # This calls the full PDF generation logic
                buffer = self.create_income_statement_pdf(summary_data, target_user)
                response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
                timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
                response["Content-Disposition"] = f'attachment; filename="income_statement_{target_user.user_id}_{timestamp}.pdf"'
                return response
            except Exception as e:
                # Log error here
                return Response({"error": f"PDF generation failed: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 4. Standard JSON Response
        return Response(summary_data)

class CurrentUserBonusSummaryView(SingleUserBonusSummaryView):
    """Endpoint for the currently authenticated user to view their own bonus summary."""
    
    # Override get to use the request.user's ID instead of the URL parameter
    def get(self, request, *args, **kwargs):
        # We replace the user_id lookup with the current authenticated user's ID
        current_user_id = request.user.user_id # Assuming user_id is an attribute
        return super().get(request, user_id=current_user_id)


class LevelUsersReport(APIView):
    pagination_class = PageNumberPagination
    pagination_class.page_size = 10
    # permission_classes = [IsAdminUser]

    def get(self, request):
        queryset = UserLevel.objects.select_related('user', 'level').prefetch_related('payments').filter(linked_user_id=request.user.user_id).order_by('-approved_at')
        
    # Query params
        email = request.query_params.get("email")
        status = request.query_params.get("status")
        user_id = request.query_params.get("user_id")
        username = request.query_params.get("username")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        limit = request.query_params.get("limit")
        from_user = request.query_params.get("from_user")
        export = request.query_params.get("export")  # 'csv', 'pdf', 'xlsx'

        # Filters
        if email:
            queryset = queryset.filter(user__email__icontains=email.lower())
        if username:
            # Annotate with a concatenated username (first_name + space + last_name)
            queryset = queryset.annotate(
                full_username=Concat('user__first_name', Value(' '), 'user__last_name', output_field=CharField())
            ).filter(
                Q(full_username__icontains=username.lower()) |
                Q(user__first_name__icontains=username.lower()) |
                Q(user__last_name__icontains=username.lower())
            )
        if status and status.lower() != "all":
            if status.lower() == "completed":
                queryset = queryset.filter(status='paid')  
            elif status.lower() == "pending":
                queryset = queryset.exclude(status='paid')
        if user_id:
            queryset = queryset.filter(user__user_id__iexact=user_id)
        if start_date:
            queryset = queryset.filter(requested_date__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(requested_date__date__lte=end_date)
        if from_user:
            queryset = queryset.annotate(
                full_from_user=Concat('user__first_name', Value(' '), 'user__last_name', output_field=CharField())
            ).filter(full_from_user__icontains=from_user.lower())

        # Search filter
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__user_id__icontains=search) |
                Q(level__name__icontains=search) |
                Q(status__icontains=search)
            )

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__gte=start_date)
            except ValueError:
                logger.error(f"Invalid start_date format: {start_date}")
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(requested_date__date__lte=end_date)
            except ValueError:
                logger.error(f"Invalid end_date format: {end_date}")
        # Limit
        if limit:
            try:
                limit = int(limit)
                queryset = queryset[:limit]
            except ValueError:
                pass

        # Export options
        if export == "csv":
            return self.export_csv(queryset, 'level_users_report')
        elif export == "pdf":
            return self.export_pdf(queryset, 'level_users_report')
        elif export == "xlsx":
            return self.export_xlsx(queryset, 'level_users_report')

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = LevelUsersSerializer(page, many=True, context={'request': request})
            return paginator.get_paginated_response(serializer.data)
        serializer = LevelUsersSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)

    def export_csv(self, queryset, filename_prefix):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        writer.writerow(['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Status', 'Level', 'Approved At', 'Total'])
        serializer = LevelUsersSerializer(queryset, many=True)
        for data in serializer.data:
            writer.writerow([
                data['from_user'],
                data['username'],
                data['from_name'],
                data['linked_username'],
                str(data['amount']),
                data['status'],
                data['level'],
                data['requested_date'] if data['requested_date'] else '',
                str(data['total'])
            ])
        return response

    def export_pdf(self, queryset, filename_prefix):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"{filename_prefix.replace('_', ' ').title()} Report", styles["Title"]))

        data = [['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Status', 'Level', 'Approved At', 'Total']]
        serializer = LevelUsersSerializer(queryset, many=True)
        for data_item in serializer.data:
            data.append([
                data_item['from_user'],
                data_item['username'],
                data_item['from_name'],
                data_item['linked_username'],
                str(data_item['amount']),
                data_item['status'],
                data_item['level'],
                data_item['requested_date'] if data_item['requested_date'] else '',
                str(data_item['total'])
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        return response

    def export_xlsx(self, queryset, filename_prefix):
        wb = Workbook()
        ws = wb.active
        ws.title = "LevelUsersReport"
        ws.append(['From User', 'Username', 'From Name', 'Linked Username', 'Amount', 'Status', 'Level', 'Approved At', 'Total'])
        serializer = LevelUsersSerializer(queryset, many=True)
        for data in serializer.data:
            ws.append([
                data['from_user'],
                data['username'],
                data['from_name'],
                data['linked_username'],
                data['amount'],
                data['status'],
                data['level'],
                data['requested_date'] if data['requested_date'] else '',
                data['total']
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response